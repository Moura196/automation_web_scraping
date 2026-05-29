"""
Output handler module for Web Scraping Project
Processes and exports scraping results to Excel files
"""

import pandas as pd
import logging
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Optional
from .logger import get_logger

logger = get_logger(__name__)


class ScrapingResult:
    """Represents a single scraping result for a product"""
    
    def __init__(self, query: str, preco: Optional[float] = None, 
                 link: Optional[str] = None, plataforma: str = "Desconhecida",
                 timestamp: Optional[str] = None):
        """
        Args:
            query: Search query used
            preco: Price in R$ (float)
            link: Product URL
            plataforma: Platform name (Google Shopping, Mercado Livre, etc)
            timestamp: When search was performed
        """
        self.query = query
        self.preco = preco
        self.link = link
        self.plataforma = plataforma
        self.timestamp = timestamp or datetime.now().isoformat()
    
    def to_dict(self) -> Dict:
        """Convert to dictionary for DataFrame"""
        return {
            'Descrição Completa': self.query,
            'Valor (R$)': self.preco,
            'Link': self.link,
            'Plataforma': self.plataforma,
            'Data da Busca': self.timestamp
        }
    
    def __repr__(self) -> str:
        return f"ScrapingResult({self.query[:30]}... | {self.preco} | {self.plataforma})"


def create_results_dataframe(results: List[ScrapingResult]) -> pd.DataFrame:
    """
    Create DataFrame from list of scraping results.
    
    Args:
        results: List of ScrapingResult objects
    
    Returns:
        DataFrame with columns:
        - Descrição Completa
        - Valor (R$)
        - Link
        - Plataforma
        - Data da Busca
    
    Example:
        >>> results = [
        ...     ScrapingResult('Luva correr', 29.90, 'https://...', 'Google Shopping'),
        ...     ScrapingResult('Luva correr', 32.50, 'https://...', 'Mercado Livre')
        ... ]
        >>> df = create_results_dataframe(results)
        >>> df.to_excel('output.xlsx')
    """
    if not results:
        logger.warning("⚠️  Nenhum resultado para exportar")
        # Return empty dataframe with correct structure
        return pd.DataFrame(columns=[
            'Descrição Completa', 'Valor (R$)', 'Link', 'Plataforma', 'Data da Busca'
        ])
    
    data = [result.to_dict() for result in results]
    df = pd.DataFrame(data)
    
    logger.info(f"✓ {len(df)} resultados processados para exportação")
    return df


def consolidate_results(multiple_results: Dict[str, List[ScrapingResult]]) -> pd.DataFrame:
    """
    Consolidate multiple scraping results into single dataframe.
    
    Groups results by query and removes duplicates.
    
    Args:
        multiple_results: Dict with query -> List[ScrapingResult]
    
    Returns:
        Consolidated DataFrame with unique results
    """
    all_results = []
    
    for query, results in multiple_results.items():
        all_results.extend(results)
    
    df = create_results_dataframe(all_results)
    
    # Remove exact duplicates (same query + same link)
    if not df.empty:
        initial_count = len(df)
        df = df.drop_duplicates(subset=['Descrição Completa', 'Link'], keep='first')
        removed = initial_count - len(df)
        
        if removed > 0:
            logger.info(f"✓ {removed} duplicatas removidas")
    
    return df


def export_results_to_excel(df: pd.DataFrame, output_path: str,
                          format_output: bool = True) -> bool:
    """
    Export results DataFrame to Excel file.
    
    Args:
        df: Results DataFrame
        output_path: Path where to save Excel file
        format_output: Whether to apply formatting (default True)
    
    Returns:
        True if successful, False otherwise
    
    Example:
        >>> success = export_results_to_excel(df, 'resultados.xlsx')
    """
    try:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Write to Excel
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Resultados', index=False)
            
            if format_output:
                # Format worksheet
                workbook = writer.book
                worksheet = writer.sheets['Resultados']
                
                # Set column widths
                column_widths = {
                    'A': 40,  # Descrição Completa
                    'B': 15,  # Valor
                    'C': 40,  # Link
                    'D': 20,  # Plataforma
                    'E': 25   # Data da Busca
                }
                
                for col, width in column_widths.items():
                    worksheet.column_dimensions[col].width = width
                
                # Format header row
                from openpyxl.styles import Font, PatternFill, Alignment
                header_fill = PatternFill(start_color='4472C4', end_color='4472C4', 
                                         fill_type='solid')
                header_font = Font(bold=True, color='FFFFFF')
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center', 
                                              wrap_text=True)
                
                # Format price column (B) as currency
                for row in worksheet.iter_rows(min_row=2, max_row=len(df)+1, 
                                             min_col=2, max_col=2):
                    for cell in row:
                        cell.number_format = 'R$ #,##0.00'
        
        logger.info(f"✓ Resultados exportados: {output_path}")
        return True
    
    except Exception as e:
        logger.error(f"❌ Erro ao exportar resultados: {str(e)}")
        return False


def get_output_filename(prefix: str = "resultados") -> str:
    """
    Generate output filename with timestamp.
    
    Args:
        prefix: Filename prefix (default "resultados")
    
    Returns:
        Filename like "resultados_29-05-2026_14-30-45.xlsx"
    """
    timestamp = datetime.now().strftime("%d-%m-%Y_%H-%M-%S")
    return f"{prefix}_{timestamp}.xlsx"


def get_default_output_path(filename: Optional[str] = None) -> Path:
    """
    Get default output directory path.
    
    Args:
        filename: Optional filename, if None generates one
    
    Returns:
        Full path to output file
    """
    output_dir = Path(__file__).parent.parent / 'output'
    
    if filename is None:
        filename = get_output_filename()
    
    return output_dir / filename
